import pygame
import sys
import openpyxl

# Initialize Pygame
pygame.init()

# Constants
WIDTH, HEIGHT = 1340, 610
BALL_RADIUS = 20
BALL_COLOR = (0, 0, 255)

# Create a window
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Ball at Coordinate")

# Load the background image
background = pygame.image.load("court.png")  # Replace with your image file
background = pygame.transform.scale(background, (WIDTH, HEIGHT))

excel_file = "shotdata.xlsx"  # Replace with your Excel file
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Define the column numbers for x and y coordinates (1-based index)
x_column = 10  # Change this to the appropriate column for x coordinates
y_column = 11  # Change this to the appropriate column for y coordinates

# Extract the x and y coordinates from the specified columns
coordinates = [(sheet.cell(row=row, column=x_column).value, sheet.cell(row=row, column=y_column).value) for row in range(1, sheet.max_row -1 )]
current_target_index = 0

# Main game loop
while True:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            pygame.quit()
            sys.exit()

    x, y = [coord * 100 for coord in coordinates[current_target_index]]
    y = HEIGHT - y

    # Draw the background image
    screen.blit(background, (0, 0))

    # Draw the ball at the specified coordinates
    pygame.draw.circle(screen, BALL_COLOR, (int(x), int(y)), BALL_RADIUS)
    current_target_index += 1
    # Update the display
    pygame.display.flip()
    pygame.time.delay(30)
